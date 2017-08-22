from vip.models import Item
from vendor.models import Store
from django.shortcuts import render
from openpyxl import load_workbook
import os
from sign_up.models import Consumer

# copy User

def copyUser(ws, user_list):

    for i, user in enumerate(user_list, start=3):
        ws.cell(row=i, column=1, value=user.username)
        ws.cell(row=i, column=2, value=user.first_name)
        ws.cell(row=i, column=3, value=user.last_name)
        ws.cell(row=i, column=4, value=user.email)
        ws.cell(row=i, column=5, value=user.phone)
        ws.cell(row=i, column=6, value=user.street)
        ws.cell(row=i, column=7, value=user.city)
        ws.cell(row=i, column=8, value=user.state)
        ws.cell(row=i, column=9, value=user.zip_code)


# copy Item
def copyItem(ws, item_list):

    for i, item in enumerate(item_list, start=3):
        ws.cell(row=i, column=1, value=item.id)
        ws.cell(row=i, column=3, value=item.name)
        ws.cell(row=i, column=4, value=item.category.name)
        ws.cell(row=i, column=5, value=item.price)
        ws.cell(row=i, column=6, value=item.store.name)
        ws.cell(row=i, column=7, value=item.description)


def copyStore(ws, store_list):

    for i, store in enumerate(store_list, start=3):
        ws.cell(row=i, column=1, value=store.id)
        ws.cell(row=i, column=2, value=store.name)
        ws.cell(row=i, column=3, value=store.street)
        ws.cell(row=i, column=4, value=store.city)
        ws.cell(row=i, column=5, value=store.state)
        ws.cell(row=i, column=6, value=store.zip_code)


# copy everything into db.xlsx
def copy (request):

    item_list = Item.objects.all()
    context = {'item_list':item_list}

    wb = load_workbook(os.path.join(os.path.dirname(__file__),'db.xlsx'))
    ws = wb['Item']
    copyItem(ws, item_list)

    consumer_list = Consumer.objects.all()
    context['user_list'] = consumer_list
    ws = wb['User']
    copyUser(ws, consumer_list)

    store_list = Store.objects.all()
    context['store_list'] = store_list
    ws = wb['Store']
    copyStore(ws, store_list)

    wb.save(os.path.join(os.path.dirname(__file__),'db.xlsx'))

    return render(request, 'copydata/index.html', context) # render looks into 'templates' folder
