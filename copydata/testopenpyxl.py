# -*- coding: utf-8 -*-
"""
Created on Mon Jul 10 12:35:19 2017

@author: rayhu
"""

from openpyxl import load_workbook


def copy ():

    wb = load_workbook('db.xlsx')
    ws = wb['VIP']
    item_list = list(range(10))
    
    for i, item in enumerate(item_list, start=3):

        ws.cell(row=i, column=3, value="item.name")
        ws.cell(row=i, column=4, value=item)
        ws.cell(row=i, column=5, value=item)
        ws.cell(row=i, column=6, value=item)
        ws.cell(row=i, column=7, value=item)
        ws.cell(row=i, column=8, value=item)
        ws.cell(row=i, column=9, value=item)
        ws.cell(row=i, column=10, value=item)

    
    wb.save('db.xlsx')


copy()